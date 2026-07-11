import json
import sys
from datetime import datetime

from openpyxl import load_workbook


SHEET_NAME = "BASE DATOS FLORIDA"


def normalize_number(value):
    number = int(value)
    if number < 0 or number > 99:
        raise ValueError("El numero debe estar entre 00 y 99.")
    return f"{number:02d}"


def normalize_date(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return datetime.fromisoformat(str(value)[:10]).strftime("%Y-%m-%d")


def read_drawings(workbook_path):
    wb = load_workbook(workbook_path)
    ws = wb[SHEET_NAME]
    drawings = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(cell is not None for cell in row):
            continue
        drawings.append(
            {
                "date": normalize_date(row[0]),
                "shift": str(row[1]).strip().upper(),
                "fijo": normalize_number(row[2]),
                "first": normalize_number(row[3]),
                "second": normalize_number(row[4]),
            }
        )

    return drawings


def append_drawing(workbook_path, drawing):
    wb = load_workbook(workbook_path)
    ws = wb[SHEET_NAME]
    date_value = normalize_date(drawing["date"])
    shift = str(drawing["shift"]).strip().upper()

    if shift not in ("T", "N"):
        raise ValueError("El turno debe ser T o N.")

    for existing in read_drawings(workbook_path):
        if existing["date"] == date_value and existing["shift"] == shift:
            raise ValueError("Ya existe una tirada para esa fecha y turno.")

    ws.append(
        [
            datetime.fromisoformat(date_value),
            shift,
            int(normalize_number(drawing["fijo"])),
            int(normalize_number(drawing["first"])),
            int(normalize_number(drawing["second"])),
        ]
    )
    wb.save(workbook_path)
    return read_drawings(workbook_path)


def main():
    payload = json.loads(sys.stdin.read())
    command = payload["command"]
    workbook_path = payload["workbookPath"]

    if command == "read":
        result = read_drawings(workbook_path)
    elif command == "append":
        result = append_drawing(workbook_path, payload["drawing"])
    else:
        raise ValueError(f"Comando no soportado: {command}")

    print(json.dumps({"ok": True, "data": result}))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(json.dumps({"ok": False, "error": str(exc)}))
        sys.exit(1)
